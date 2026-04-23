"""Parsers: CSV / XLSX / YAML bytes → ModelErrorCodes."""

from __future__ import annotations

import io
from typing import Any

import yaml
from pydantic import ValidationError

from .schema import SUPPORTED_VENDORS, ErrorCode, ModelErrorCodes

REQUIRED_COLUMNS = ("code", "description", "severity")
OPTIONAL_COLUMNS = ("component", "notes")
_ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


class ParseError(ValueError):
    """Raised when uploaded file cannot be parsed into ModelErrorCodes."""


def _normalize_col(name: str) -> str:
    return name.strip().lower().replace(" ", "_")


def _coerce_severity(value: Any) -> str:
    s = str(value or "").strip()
    lower = s.lower()
    for canonical in ("Critical", "High", "Medium", "Low", "Info"):
        if lower == canonical.lower():
            return canonical
    # common synonyms
    synonyms = {"major": "High", "minor": "Low", "warning": "Medium"}
    if lower in synonyms:
        return synonyms[lower]
    raise ParseError(
        f"Недопустимое значение severity='{value}'. "
        f"Разрешено: Critical/High/Medium/Low/Info."
    )


def _rows_to_model(
    rows: list[dict[str, Any]],
    *,
    vendor: str,
    model: str,
) -> ModelErrorCodes:
    if not rows:
        raise ParseError("Файл не содержит строк с данными.")
    codes: dict[str, ErrorCode] = {}
    errors: list[str] = []
    for idx, row in enumerate(rows, start=2):  # row 1 — header
        try:
            code = str(row.get("code") or "").strip()
            if not code:
                continue
            description = str(row.get("description") or "").strip()
            if not description:
                errors.append(f"Строка {idx}: пустое описание для кода {code}")
                continue
            severity = _coerce_severity(row.get("severity"))
            component = str(row.get("component") or "").strip()
            notes = str(row.get("notes") or "").strip()
            codes[code] = ErrorCode(
                description=description,
                severity=severity,  # type: ignore[arg-type]
                component=component,
                notes=notes,
            )
        except (ParseError, ValidationError) as exc:
            errors.append(f"Строка {idx}: {exc}")
    if errors and not codes:
        raise ParseError("Ни одной валидной строки: " + "; ".join(errors[:5]))
    if not codes:
        raise ParseError("После парсинга не осталось ни одного валидного кода.")
    try:
        return ModelErrorCodes(vendor=vendor, model=model, codes=codes)
    except ValidationError as exc:
        raise ParseError(f"Ошибка валидации: {exc}") from exc


def _validate_headers(headers: list[str]) -> None:
    normalized = {_normalize_col(h) for h in headers}
    missing = [c for c in REQUIRED_COLUMNS if c not in normalized]
    if missing:
        raise ParseError(
            f"Отсутствуют обязательные колонки: {', '.join(missing)}. "
            f"Нужны как минимум: {', '.join(REQUIRED_COLUMNS)}."
        )


def parse_csv(data: bytes | str, *, vendor: str, model: str) -> ModelErrorCodes:
    """Parse CSV into ModelErrorCodes. Accepts bytes or str."""
    import csv

    if isinstance(data, bytes):
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("cp1251", errors="replace")
    else:
        text = data

    # Sniff delimiter
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # fallback

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if reader.fieldnames is None:
        raise ParseError("CSV не содержит заголовков.")
    _validate_headers(list(reader.fieldnames))

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row = {
            _normalize_col(k): v
            for k, v in raw_row.items()
            if k is not None and _normalize_col(k) in _ALL_COLUMNS
        }
        rows.append(row)
    return _rows_to_model(rows, vendor=vendor, model=model)


def parse_xlsx(data: bytes, *, vendor: str, model: str) -> ModelErrorCodes:
    """Parse XLSX first sheet into ModelErrorCodes."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ParseError("openpyxl не установлен — XLSX не поддерживается.") from exc

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ParseError(f"Не удалось открыть XLSX: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise ParseError("XLSX не содержит листов.")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ParseError("XLSX пустой.") from exc

    headers = [_normalize_col(str(h) if h is not None else "") for h in header_row]
    _validate_headers(headers)

    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        if values is None or all(v is None or v == "" for v in values):
            continue
        row = {
            headers[i]: values[i]
            for i in range(min(len(headers), len(values)))
            if headers[i] in _ALL_COLUMNS
        }
        rows.append(row)
    return _rows_to_model(rows, vendor=vendor, model=model)


def parse_yaml(data: bytes | str, *, vendor: str, model: str) -> ModelErrorCodes:
    """Parse YAML bytes/string into ModelErrorCodes.

    Expected structure:
      vendor: Xerox
      model: B8090
      codes:
        "09-605-00":
          description: ...
          severity: Critical
          component: fuser
    """
    if isinstance(data, bytes):
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("utf-8", errors="replace")
    else:
        text = data

    try:
        parsed = yaml.safe_load(text) or {}
    except yaml.YAMLError as exc:
        raise ParseError(f"YAML parse error: {exc}") from exc
    if not isinstance(parsed, dict):
        raise ParseError("YAML: ожидался объект на верхнем уровне.")

    codes_raw = parsed.get("codes") or {}
    if not isinstance(codes_raw, dict):
        raise ParseError("YAML: поле 'codes' должно быть объектом.")

    codes: dict[str, ErrorCode] = {}
    errors: list[str] = []
    for code, info in codes_raw.items():
        if not isinstance(info, dict):
            errors.append(f"Код {code}: значение не является объектом.")
            continue
        try:
            sev = _coerce_severity(info.get("severity"))
            codes[str(code)] = ErrorCode(
                description=str(info.get("description") or "").strip(),
                severity=sev,  # type: ignore[arg-type]
                component=str(info.get("component") or "").strip(),
                notes=str(info.get("notes") or "").strip(),
            )
        except (ParseError, ValidationError) as exc:
            errors.append(f"Код {code}: {exc}")
    if not codes:
        raise ParseError("После парсинга не осталось ни одного валидного кода.")

    try:
        return ModelErrorCodes(
            vendor=parsed.get("vendor") or vendor,
            model=parsed.get("model") or model,
            codes=codes,
        )
    except ValidationError as exc:
        raise ParseError(f"Ошибка валидации: {exc}") from exc


def parse_file(
    filename: str,
    data: bytes,
    *,
    vendor: str,
    model: str,
) -> ModelErrorCodes:
    """Dispatch by extension."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_csv(data, vendor=vendor, model=model)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(data, vendor=vendor, model=model)
    if name.endswith(".yaml") or name.endswith(".yml"):
        return parse_yaml(data, vendor=vendor, model=model)
    raise ParseError(
        f"Неподдерживаемое расширение файла: {filename}. "
        f"Разрешены: .csv, .xlsx, .yaml"
    )
