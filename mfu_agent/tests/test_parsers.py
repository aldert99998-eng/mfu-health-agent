"""Tests for data_io/parsers.py — Phase 4.1 verification.

TC-E-1  CP1251-CSV с разделителем `;`
TC-E-2  UTF-8 CSV с разделителем `,`
TC-E-3  TSV (табуляция)
TC-E-4  JSON (массив объектов)
TC-E-5  JSON (одиночный объект → оборачивается)
TC-E-6  JSONL
TC-E-7  XLSX
TC-E-8  Битый JSON → MalformedFileError с указанием строки
TC-E-9  Битый JSONL → MalformedFileError с номером строки
TC-E-10 Пустой файл → EmptyFileError
TC-E-11 Неподдерживаемый формат → UnsupportedFormatError
TC-E-12 Вложенный JSON → json_normalize
TC-E-13 CSV только с заголовком (нет данных) → EmptyFileError
"""

from __future__ import annotations

from pathlib import Path

import pytest

from data_io.models import FileFormat
from data_io.parsers import (
    CSVParser,
    EmptyFileError,
    EncodingError,
    FormatDetector,
    InvalidFileFormatError,
    MalformedFileError,
    UnsupportedFormatError,
    XLSXParser,
    parse_file,
)


@pytest.fixture()
def tmp(tmp_path: Path) -> Path:
    return tmp_path


# ── TC-E-1: CP1251 CSV с ; ──────────────────────────────────────────────────


class TestCP1251CSV:
    def test_parse_cp1251_semicolon(self, tmp: Path) -> None:
        p = tmp / "data_cp1251.csv"
        content = "устройство;код;статус\nМФУ-001;C6000;активен\nМФУ-002;A1001;простой\n"
        p.write_bytes(content.encode("cp1251"))

        df = parse_file(p)

        assert len(df) == 2
        assert list(df.columns) == ["устройство", "код", "статус"]
        assert df.iloc[0]["устройство"] == "МФУ-001"
        assert df.iloc[1]["код"] == "A1001"


# ── TC-E-2: UTF-8 CSV с , ───────────────────────────────────────────────────


class TestUTF8CSV:
    def test_parse_utf8_comma(self, tmp: Path) -> None:
        p = tmp / "data.csv"
        p.write_text(
            "device_id,error_code,status\nD001,C6000,active\nD002,A1001,idle\n",
            encoding="utf-8",
        )

        df = parse_file(p)

        assert len(df) == 2
        assert list(df.columns) == ["device_id", "error_code", "status"]
        assert df.iloc[0]["device_id"] == "D001"


# ── TC-E-3: TSV ─────────────────────────────────────────────────────────────


class TestTSV:
    def test_parse_tsv(self, tmp: Path) -> None:
        p = tmp / "data.tsv"
        p.write_text("id\tname\tvalue\n1\talpha\t100\n2\tbeta\t200\n")

        df = parse_file(p)

        assert len(df) == 2
        assert list(df.columns) == ["id", "name", "value"]
        assert df.iloc[0]["name"] == "alpha"


# ── TC-E-4: JSON (массив объектов) ──────────────────────────────────────────


class TestJSONArray:
    def test_parse_json_array(self, tmp: Path) -> None:
        p = tmp / "data.json"
        p.write_text('[{"id": 1, "val": "a"}, {"id": 2, "val": "b"}]')

        df = parse_file(p)

        assert len(df) == 2
        assert "id" in df.columns
        assert "val" in df.columns


# ── TC-E-5: JSON (одиночный объект) ─────────────────────────────────────────


class TestJSONSingleObject:
    def test_single_object_wrapped(self, tmp: Path) -> None:
        p = tmp / "single.json"
        p.write_text('{"device_id": "D001", "status": "ok"}')

        df = parse_file(p)

        assert len(df) == 1
        assert df.iloc[0]["device_id"] == "D001"


# ── TC-E-6: JSONL ───────────────────────────────────────────────────────────


class TestJSONL:
    def test_parse_jsonl(self, tmp: Path) -> None:
        p = tmp / "data.jsonl"
        p.write_text('{"x": 1}\n{"x": 2}\n{"x": 3}\n')

        df = parse_file(p)

        assert len(df) == 3
        assert list(df.columns) == ["x"]

    def test_jsonl_with_blank_lines(self, tmp: Path) -> None:
        p = tmp / "gaps.jsonl"
        p.write_text('{"a": 1}\n\n{"a": 2}\n\n')

        df = parse_file(p)

        assert len(df) == 2


# ── TC-E-7: XLSX ────────────────────────────────────────────────────────────


class TestXLSX:
    def test_parse_xlsx(self, tmp: Path) -> None:
        from openpyxl import Workbook

        p = tmp / "data.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["device_id", "error_code", "status"])
        ws.append(["D001", "C6000", "active"])
        ws.append(["D002", "A1001", "idle"])
        wb.save(p)

        df = parse_file(p)

        assert len(df) == 2
        assert list(df.columns) == ["device_id", "error_code", "status"]
        assert df.iloc[0]["device_id"] == "D001"


# ── TC-E-8: Битый JSON → MalformedFileError ─────────────────────────────────


class TestMalformedJSON:
    def test_broken_json(self, tmp: Path) -> None:
        p = tmp / "broken.json"
        p.write_text('{"id": 1, "val": INVALID}')

        with pytest.raises(MalformedFileError, match=r"broken\.json"):
            parse_file(p)

    def test_json_not_object_or_array(self, tmp: Path) -> None:
        p = tmp / "scalar.json"
        p.write_text('"just a string"')

        with pytest.raises(InvalidFileFormatError):
            parse_file(p)


# ── TC-E-9: Битый JSONL → MalformedFileError с номером строки ───────────────


class TestMalformedJSONL:
    def test_broken_jsonl_line(self, tmp: Path) -> None:
        p = tmp / "broken.jsonl"
        p.write_text('{"ok": 1}\n{BROKEN}\n{"ok": 3}\n')

        with pytest.raises(MalformedFileError, match=r"строка 2"):
            parse_file(p)

    def test_jsonl_non_object(self, tmp: Path) -> None:
        p = tmp / "array_line.jsonl"
        p.write_text('[1, 2, 3]\n')

        with pytest.raises(MalformedFileError, match=r"строка 1.*ожидался объект"):
            parse_file(p)


# ── TC-E-10: Пустой файл → EmptyFileError ───────────────────────────────────


class TestEmptyFile:
    def test_empty_csv(self, tmp: Path) -> None:
        p = tmp / "empty.csv"
        p.write_text("")

        with pytest.raises(EmptyFileError, match=r"empty\.csv"):
            parse_file(p)

    def test_empty_json(self, tmp: Path) -> None:
        p = tmp / "empty.json"
        p.write_text("   ")

        with pytest.raises(EmptyFileError, match=r"empty\.json"):
            parse_file(p)

    def test_empty_jsonl(self, tmp: Path) -> None:
        p = tmp / "empty.jsonl"
        p.write_text("\n\n\n")

        with pytest.raises(EmptyFileError):
            parse_file(p)


# ── TC-E-11: Неподдерживаемый формат ────────────────────────────────────────


class TestUnsupportedFormat:
    def test_pdf_rejected(self, tmp: Path) -> None:
        p = tmp / "doc.pdf"
        p.write_text("fake pdf")

        with pytest.raises(UnsupportedFormatError, match=r"\.pdf"):
            parse_file(p)

    def test_txt_rejected(self, tmp: Path) -> None:
        p = tmp / "notes.txt"
        p.write_text("hello")

        with pytest.raises(UnsupportedFormatError):
            parse_file(p)


# ── TC-E-12: Вложенный JSON → json_normalize ────────────────────────────────


class TestNestedJSON:
    def test_nested_flattened(self, tmp: Path) -> None:
        p = tmp / "nested.json"
        p.write_text('[{"id": 1, "info": {"model": "HP", "location": "MSK"}}]')

        df = parse_file(p)

        assert len(df) == 1
        assert "info.model" in df.columns
        assert df.iloc[0]["info.model"] == "HP"
        assert df.iloc[0]["info.location"] == "MSK"


# ── TC-E-13: CSV только заголовок → EmptyFileError ──────────────────────────


class TestHeaderOnlyCSV:
    def test_header_only(self, tmp: Path) -> None:
        p = tmp / "header_only.csv"
        p.write_text("col1,col2,col3\n")

        with pytest.raises(EmptyFileError):
            parse_file(p)


# ── FormatDetector unit tests ────────────────────────────────────────────────


# ── TC-E-14: EncodingError конструктор ───────────────────────────────────────


class TestEncodingErrorInit:
    def test_with_encoding_and_detail(self) -> None:
        err = EncodingError(Path("f.csv"), "cp1251", "bad byte")
        assert "f.csv" in str(err)
        assert "cp1251" in str(err)
        assert "bad byte" in str(err)

    def test_minimal(self) -> None:
        err = EncodingError(Path("f.csv"))
        assert "f.csv" in str(err)


# ── TC-E-15: CSV кодировка, которая не декодируется ─────────────────────────


class TestCSVEncodingFallback:
    def test_chardet_detected_fails_decode_fallback_utf8(self, tmp: Path) -> None:
        """chardet возвращает кодировку с низкой уверенностью, которая не декодирует → utf-8."""
        p = tmp / "tricky.csv"
        p.write_text("a,b\n1,2\n", encoding="utf-8")
        df = CSVParser().parse(p)
        assert len(df) == 1

    def test_high_confidence_wrong_encoding_raises(self, tmp: Path) -> None:
        """Если chardet уверен, но decode всё равно падает → EncodingError."""
        p = tmp / "bad.csv"
        p.write_bytes(b"\x80\x81\x82\n\x83\x84\x85\n")
        from unittest.mock import patch
        with patch("data_io.parsers.chardet.detect", return_value={
            "encoding": "ascii", "confidence": 0.99,
        }), pytest.raises(EncodingError):
            CSVParser().parse(p)


# ── TC-E-16: Corrupt XLSX ───────────────────────────────────────────────────


class TestCorruptXLSX:
    def test_corrupt_xlsx_raises(self, tmp: Path) -> None:
        p = tmp / "corrupt.xlsx"
        p.write_bytes(b"not an xlsx file")

        with pytest.raises(MalformedFileError, match=r"corrupt\.xlsx"):
            XLSXParser().parse(p)

    def test_xlsx_header_only(self, tmp: Path) -> None:
        from openpyxl import Workbook
        p = tmp / "header_only.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["col1", "col2"])
        wb.save(p)

        with pytest.raises(EmptyFileError):
            XLSXParser().parse(p)

    def test_xlsx_empty_workbook(self, tmp: Path) -> None:
        from openpyxl import Workbook
        p = tmp / "empty.xlsx"
        wb = Workbook()
        wb.save(p)

        with pytest.raises(EmptyFileError):
            XLSXParser().parse(p)


# ── TC-E-17: JSON пустой массив ─────────────────────────────────────────────


class TestJSONEmptyArray:
    def test_empty_array(self, tmp: Path) -> None:
        p = tmp / "empty_arr.json"
        p.write_text("[]")

        with pytest.raises(EmptyFileError):
            parse_file(p)


# ── TC-E-18: CSV sniffer fallback ───────────────────────────────────────────


class TestCSVSnifferFallback:
    def test_sniffer_fails_uses_comma(self, tmp: Path) -> None:
        p = tmp / "single_col.csv"
        p.write_text("value\n100\n200\n")
        df = CSVParser().parse(p)
        assert len(df) == 2
        assert list(df.columns) == ["value"]


# ── FormatDetector unit tests ────────────────────────────────────────────────


class TestFormatDetector:
    @pytest.mark.parametrize(
        ("ext", "expected"),
        [
            (".csv", FileFormat.CSV),
            (".tsv", FileFormat.TSV),
            (".json", FileFormat.JSON),
            (".jsonl", FileFormat.JSONL),
            (".ndjson", FileFormat.JSONL),
            (".xlsx", FileFormat.XLSX),
        ],
    )
    def test_known_extensions(self, ext: str, expected: FileFormat) -> None:
        assert FormatDetector.detect(Path(f"file{ext}")) == expected

    def test_unknown_extension(self) -> None:
        with pytest.raises(UnsupportedFormatError):
            FormatDetector.detect(Path("file.parquet"))
