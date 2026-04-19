"""File parsers for ingestion pipeline — Track E core.

FormatDetector  → определяет формат файла по расширению и содержимому.
FileParser      → абстрактный базовый класс.
CSVParser       → CSV / TSV с автоопределением кодировки и разделителя.
JSONParser      → JSON / JSONL (массив или построчно).
XLSXParser      → Excel через openpyxl (read-only).
parse_file()    → диспетчер: путь → DataFrame.
"""

from __future__ import annotations

import csv
import io
import json
import re
import logging
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

import chardet
import pandas as pd
from openpyxl import load_workbook

from data_io.models import FileFormat

logger = logging.getLogger(__name__)

# ── Error classes ─────────────────────────────────────────────────────────────


class UnsupportedFormatError(Exception):
    """Формат файла не поддерживается."""

    def __init__(self, path: Path, detail: str = "") -> None:
        self.path = path
        msg = f"Формат файла не поддерживается: {path.name}"
        if detail:
            msg += f" ({detail})"
        super().__init__(msg)


class InvalidFileFormatError(Exception):
    """Расширение файла не соответствует содержимому."""

    def __init__(self, path: Path, expected: str, actual: str) -> None:
        self.path = path
        self.expected = expected
        self.actual = actual
        super().__init__(
            f"Содержимое файла {path.name} не соответствует расширению: "
            f"ожидается {expected}, обнаружено {actual}"
        )


class EncodingError(Exception):
    """Не удалось определить или декодировать кодировку файла."""

    def __init__(self, path: Path, encoding: str = "", detail: str = "") -> None:
        self.path = path
        self.encoding = encoding
        msg = f"Ошибка кодировки файла: {path.name}"
        if encoding:
            msg += f" [{encoding}]"
        if detail:
            msg += f" — {detail}"
        super().__init__(msg)


class EmptyFileError(Exception):
    """Файл пуст или не содержит данных."""

    def __init__(self, path: Path) -> None:
        self.path = path
        super().__init__(f"Файл пуст или не содержит данных: {path.name}")


class MalformedFileError(Exception):
    """Файл повреждён или имеет некорректную структуру."""

    def __init__(self, path: Path, detail: str = "") -> None:
        self.path = path
        msg = f"Некорректная структура файла: {path.name}"
        if detail:
            msg += f" — {detail}"
        super().__init__(msg)


# ── Magic bytes for content-based validation (E-005, E-006) ──────────────────

_XLSX_MAGIC = b"PK\x03\x04"
_JSON_START = frozenset(b"[{")


def _detect_format_by_content(path: Path) -> FileFormat | None:
    """Detect file format by inspecting content/magic bytes."""
    try:
        head = path.read_bytes()[:8192]
    except OSError:
        return None

    if not head:
        return None

    if head[:4] == _XLSX_MAGIC:
        return FileFormat.XLSX

    try:
        text = head.decode("utf-8", errors="ignore").lstrip()
    except Exception:
        return None

    if not text:
        return None

    first_char = text[0] if text else ""

    if first_char in ("{", "["):
        if "\n" in text:
            lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
            if len(lines) >= 2 and all(ln.startswith("{") for ln in lines[:5]):
                return FileFormat.JSONL
        return FileFormat.JSON

    if "\t" in text.split("\n")[0] and "," not in text.split("\n")[0]:
        return FileFormat.TSV

    if "," in text.split("\n")[0] or ";" in text.split("\n")[0]:
        return FileFormat.CSV

    return None


def _validate_magic_bytes(path: Path, declared: FileFormat) -> None:
    """Validate that file content matches the declared format (E-006)."""
    if declared == FileFormat.XLSX:
        try:
            magic = path.read_bytes()[:4]
        except OSError:
            return
        if magic and magic != _XLSX_MAGIC:
            raise InvalidFileFormatError(path, "XLSX (ZIP/PK)", "другой формат")

    elif declared in (FileFormat.JSON, FileFormat.JSONL):
        try:
            head = path.read_bytes()[:256]
        except OSError:
            return
        if not head:
            return
        text = head.decode("utf-8", errors="ignore").lstrip()
        if text and ord(text[0]) not in _JSON_START:
            raise InvalidFileFormatError(path, "JSON", "не JSON-содержимое")


# ── FormatDetector ────────────────────────────────────────────────────────────

_EXT_MAP: dict[str, FileFormat] = {
    ".csv": FileFormat.CSV,
    ".tsv": FileFormat.TSV,
    ".json": FileFormat.JSON,
    ".jsonl": FileFormat.JSONL,
    ".ndjson": FileFormat.JSONL,
    ".xlsx": FileFormat.XLSX,
}


class FormatDetector:
    """Определяет FileFormat по расширению и содержимому файла."""

    @staticmethod
    def detect(path: Path) -> FileFormat:
        ext = path.suffix.lower()
        fmt = _EXT_MAP.get(ext)

        if fmt is not None:
            _validate_magic_bytes(path, fmt)
            logger.info("Формат файла %s → %s", path.name, fmt.value)
            return fmt

        content_fmt = _detect_format_by_content(path)
        if content_fmt is not None:
            logger.info(
                "Формат файла %s → %s (определён по содержимому, расширение '%s')",
                path.name, content_fmt.value, ext,
            )
            return content_fmt

        raise UnsupportedFormatError(path, f"расширение '{ext}'")


# ── Abstract FileParser ──────────────────────────────────────────────────────


class FileParser(ABC):
    """Абстрактный парсер: path → DataFrame."""

    @abstractmethod
    def parse(self, path: Path) -> pd.DataFrame:
        ...


# ── CSVParser ─────────────────────────────────────────────────────────────────


class CSVParser(FileParser):
    """Парсер CSV / TSV с автоопределением кодировки и разделителя."""

    def parse(self, path: Path) -> pd.DataFrame:
        raw = path.read_bytes()
        if not raw.strip():
            raise EmptyFileError(path)

        encoding = self._detect_encoding(path, raw)
        try:
            text = raw.decode(encoding)
        except (UnicodeDecodeError, LookupError) as exc:
            raise EncodingError(path, encoding, str(exc)) from exc

        text = self._skip_preamble(text)

        delimiter = self._detect_delimiter(path, text)
        logger.info(
            "CSV: файл=%s, кодировка=%s, разделитель=%r",
            path.name, encoding, delimiter,
        )

        try:
            df = pd.read_csv(io.StringIO(text), sep=delimiter, dtype=str)
        except Exception as exc:
            raise MalformedFileError(path, str(exc)) from exc

        if df.empty:
            raise EmptyFileError(path)

        df.columns = df.columns.str.strip()
        logger.info("CSV: %d строк, %d столбцов", len(df), len(df.columns))
        return df

    @staticmethod
    def _detect_encoding(path: Path, raw: bytes) -> str:
        result = chardet.detect(raw)
        detected = result.get("encoding")
        confidence = result.get("confidence", 0.0)

        if not detected:
            return "utf-8"

        if confidence < 0.5:
            logger.info(
                "Низкая уверенность кодировки (%.0f%%) для %s, пробуем %s → utf-8",
                confidence * 100, path.name, detected,
            )
            try:
                raw.decode(detected)
                return detected
            except (UnicodeDecodeError, LookupError):
                return "utf-8"

        return detected

    @staticmethod
    def _skip_preamble(text: str) -> str:
        """Skip BOM, SQL comments/queries, and other non-data preamble."""
        if text.startswith("\ufeff"):
            text = text[1:]

        _SQL_KEYWORDS = {"SELECT", "UNION", "FROM", "WHERE", "INNER", "LEFT", "JOIN", "INSERT", "CREATE"}

        lines = text.split("\n")
        start = 0
        for i, line in enumerate(lines):
            stripped = line.strip()
            if not stripped:
                start = i + 1
                continue
            if stripped.startswith("--"):
                start = i + 1
                continue
            upper = stripped.upper()
            tokens = set(re.split(r"[\s(,;]+", upper[:200]))
            if tokens & _SQL_KEYWORDS:
                start = i + 1
                continue
            break
        if start > 0:
            logger.info("CSV: пропущено %d строк преамбулы (SQL/комментарии)", start)
            return "\n".join(lines[start:])
        return text

    @staticmethod
    def _detect_delimiter(path: Path, text: str) -> str:
        if path.suffix.lower() == ".tsv":
            return "\t"
        try:
            sample = text[:8192]
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            return dialect.delimiter
        except csv.Error:
            return ","


# ── JSONParser ────────────────────────────────────────────────────────────────


class JSONParser(FileParser):
    """Парсер JSON (массив объектов) и JSONL (по одному объекту на строку)."""

    def parse(self, path: Path) -> pd.DataFrame:
        raw = path.read_bytes()
        if not raw.strip():
            raise EmptyFileError(path)

        text = raw.decode("utf-8")

        if path.suffix.lower() in (".jsonl", ".ndjson"):
            records = self._parse_jsonl(path, text)
        else:
            records = self._parse_json(path, text)

        if not records:
            raise EmptyFileError(path)

        df = pd.json_normalize(records)
        df = df.astype(str)
        logger.info("JSON: %d записей, %d столбцов", len(df), len(df.columns))
        return df

    @staticmethod
    def _parse_json(path: Path, text: str) -> list[dict[str, Any]]:
        try:
            data = json.loads(text)
        except json.JSONDecodeError as exc:
            raise MalformedFileError(path, str(exc)) from exc

        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            if len(data) == 1:
                val = next(iter(data.values()))
                if isinstance(val, list) and val and isinstance(val[0], dict):
                    logger.info("JSON: извлечён массив из обёртки (1 ключ → массив)")
                    return val
            return [data]
        raise MalformedFileError(path, "ожидался массив или объект")

    @staticmethod
    def _parse_jsonl(path: Path, text: str) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for i, line in enumerate(text.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as exc:
                raise MalformedFileError(
                    path, f"строка {i}: {exc}"
                ) from exc
            if not isinstance(obj, dict):
                raise MalformedFileError(
                    path, f"строка {i}: ожидался объект, получен {type(obj).__name__}"
                )
            records.append(obj)
        return records


# ── XLSXParser ────────────────────────────────────────────────────────────────


class XLSXParser(FileParser):
    """Парсер Excel (.xlsx) через openpyxl в режиме read-only.

    Reads all non-empty sheets and concatenates them.
    """

    def parse(self, path: Path) -> pd.DataFrame:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise MalformedFileError(path, str(exc)) from exc

        try:
            frames: list[pd.DataFrame] = []
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 2:
                    continue
                header = [
                    str(c).strip() if c is not None else f"col_{i}"
                    for i, c in enumerate(rows[0])
                ]
                data = rows[1:]
                sheet_df = pd.DataFrame(data, columns=header).astype(str)
                if not sheet_df.empty:
                    logger.info(
                        "XLSX лист '%s': %d строк, %d столбцов",
                        ws.title, len(sheet_df), len(sheet_df.columns),
                    )
                    frames.append(sheet_df)
        finally:
            wb.close()

        if not frames:
            raise EmptyFileError(path)

        if len(frames) == 1:
            df = frames[0]
        else:
            df = pd.concat(frames, ignore_index=True)
            logger.info("XLSX: объединено %d листов, итого %d строк", len(frames), len(df))

        return df


# ── Dispatcher ────────────────────────────────────────────────────────────────

_PARSER_MAP: dict[FileFormat, type[FileParser]] = {
    FileFormat.CSV: CSVParser,
    FileFormat.TSV: CSVParser,
    FileFormat.JSON: JSONParser,
    FileFormat.JSONL: JSONParser,
    FileFormat.XLSX: XLSXParser,
}


def _check_duplicate_headers(df: pd.DataFrame, path: Path) -> pd.DataFrame:
    """Detect and rename duplicate column headers (E-014)."""
    cols = list(df.columns)
    seen: dict[str, int] = {}
    renamed = False
    new_cols: list[str] = []
    for col in cols:
        if col in seen:
            seen[col] += 1
            new_name = f"{col}_{seen[col]}"
            new_cols.append(new_name)
            renamed = True
        else:
            seen[col] = 0
            new_cols.append(col)
    if renamed:
        dups = [c for c, cnt in seen.items() if cnt > 0]
        logger.warning(
            "Дублирующиеся заголовки в %s: %s — переименованы автоматически",
            path.name, ", ".join(dups),
        )
        df.columns = pd.Index(new_cols)
    return df


def parse_file(path: Path) -> pd.DataFrame:
    """Диспетчер: определяет формат и парсит файл в DataFrame.

    Raises:
        UnsupportedFormatError, EncodingError, EmptyFileError,
        MalformedFileError, InvalidFileFormatError
    """
    fmt = FormatDetector.detect(path)
    parser_cls = _PARSER_MAP.get(fmt)
    if parser_cls is None:
        raise UnsupportedFormatError(path)

    parser = parser_cls()
    logger.info("Парсинг файла %s (формат: %s)", path.name, fmt.value)
    df = parser.parse(path)
    df = _check_duplicate_headers(df, path)
    return df
